"""
Módulo de Geração de Relatórios
Gera gráficos e arquivos Excel formatados
"""

import os
import logging
import pandas as pd
import matplotlib
matplotlib.use('Agg')  # Backend sem display
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

logger = logging.getLogger(__name__)


class ReportGenerator:
    """
    Gera relatórios visuais e arquivos Excel formatados
    """

    def __init__(self, output_dir: str = "output"):
        """
        Inicializa o gerador de relatórios

        Args:
            output_dir: Diretório para salvar arquivos gerados
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def gerar_grafico_ranking(self, df_ranking: pd.DataFrame, titulo: str = "Ranking de Origens", cor: str = '#ffc107') -> str:
        """
        Gera gráfico de barras horizontais do ranking de origens

        Args:
            df_ranking: DataFrame com ranking (Origem, Total Negadas ou Total N2)
            titulo: Título do gráfico
            cor: Cor das barras

        Returns:
            Caminho do arquivo de imagem gerado
        """
        try:
            if df_ranking is None or len(df_ranking) == 0:
                logger.warning("DataFrame vazio, não foi possível gerar gráfico")
                return None

            # Configurar figura
            fig, ax = plt.subplots(figsize=(12, 8))

            # Preparar dados - detectar coluna de valores
            origens = df_ranking['Origem'].values

            # Determinar qual coluna de valores usar
            if 'Total Negadas' in df_ranking.columns:
                valores = df_ranking['Total Negadas'].values
                label = 'Total Negadas'
            elif 'Total N2' in df_ranking.columns:
                valores = df_ranking['Total N2'].values
                label = 'Total N2'
            else:
                logger.error("Coluna de valores não encontrada no DataFrame")
                return None

            # Posições das barras
            y_pos = range(len(origens))

            # Criar barras
            bars = ax.barh(y_pos, valores, color=cor, alpha=0.8, edgecolor='black', linewidth=0.5)

            # Adicionar valores nas barras
            for bar in bars:
                width = bar.get_width()
                if width > 0:
                    ax.text(width, bar.get_y() + bar.get_height()/2,
                           f' {int(width)}', va='center', ha='left', fontsize=10, fontweight='bold')

            # Configurar eixos
            ax.set_yticks(y_pos)
            ax.set_yticklabels(origens, fontsize=10)
            ax.invert_yaxis()  # Top = primeira origem
            ax.set_xlabel('Quantidade de Transações', fontsize=12, fontweight='bold')
            ax.set_title(titulo, fontsize=14, fontweight='bold', pad=20)

            # Grid
            ax.grid(axis='x', alpha=0.3, linestyle='--')

            # Layout ajustado
            plt.tight_layout()

            # Salvar
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            suffix = "negadas" if label == 'Total Negadas' else "n2"
            filename = f"ranking_{suffix}_{timestamp}.png"
            filepath = os.path.join(self.output_dir, filename)

            plt.savefig(filepath, dpi=150, bbox_inches='tight', facecolor='white')
            plt.close()

            logger.info(f"Gráfico gerado: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Erro ao gerar gráfico: {e}")
            return None

    def gerar_excel_completo(self,
                            tabela_resumo: pd.DataFrame,
                            tabela_codigos: pd.DataFrame,
                            ranking_negadas: pd.DataFrame,
                            ranking_n2: pd.DataFrame,
                            tabela_negadas: pd.DataFrame) -> str:
        """
        Gera arquivo Excel completo e formatado

        Args:
            tabela_resumo: DataFrame com resumo geral
            tabela_codigos: DataFrame com códigos de resposta
            ranking_negadas: DataFrame com ranking de todas as negadas
            ranking_n2: DataFrame com ranking específico de N2
            tabela_negadas: DataFrame com transações negadas

        Returns:
            Caminho do arquivo Excel gerado
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"relatorio_recargas_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)

            # Criar Excel com múltiplas abas
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Aba 1: Resumo Geral
                if tabela_resumo is not None and len(tabela_resumo) > 0:
                    tabela_resumo.to_excel(writer, sheet_name='Resumo Geral', index=False)

                # Aba 2: Ranking Negadas
                if ranking_negadas is not None and len(ranking_negadas) > 0:
                    ranking_negadas.to_excel(writer, sheet_name='Ranking Negadas', index=True, index_label='#')

                # Aba 3: Ranking N2
                if ranking_n2 is not None and len(ranking_n2) > 0:
                    ranking_n2.to_excel(writer, sheet_name='Ranking N2', index=True, index_label='#')

                # Aba 4: Códigos de Resposta
                if tabela_codigos is not None and len(tabela_codigos) > 0:
                    tabela_codigos.to_excel(writer, sheet_name='Códigos de Resposta', index=False)

                # Aba 5: Recargas Negadas
                if tabela_negadas is not None and len(tabela_negadas) > 0:
                    tabela_negadas.to_excel(writer, sheet_name='Recargas Negadas', index=False)

            # Formatar Excel
            self._formatar_excel(filepath)

            logger.info(f"Excel gerado: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Erro ao gerar Excel: {e}")
            return None

    def _formatar_excel(self, filepath: str):
        """
        Aplica formatação profissional ao Excel

        Args:
            filepath: Caminho do arquivo Excel
        """
        try:
            wb = load_workbook(filepath)

            # Estilos
            header_fill = PatternFill(start_color="4a5568", end_color="4a5568", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Formatar cada aba
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Formatar cabeçalho
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border

                # Ajustar largura das colunas
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                # Aplicar bordas e alinhamento
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # Congelar primeira linha
                ws.freeze_panes = 'A2'

            wb.save(filepath)
            logger.info(f"Excel formatado com sucesso: {filepath}")

        except Exception as e:
            logger.error(f"Erro ao formatar Excel: {e}")

    def limpar_arquivos_antigos(self, dias: int = 7):
        """
        Remove arquivos gerados há mais de X dias

        Args:
            dias: Número de dias para manter arquivos
        """
        try:
            agora = datetime.now()
            count = 0

            for filename in os.listdir(self.output_dir):
                filepath = os.path.join(self.output_dir, filename)

                if not os.path.isfile(filepath):
                    continue

                # Verificar idade do arquivo
                file_time = datetime.fromtimestamp(os.path.getmtime(filepath))
                idade = (agora - file_time).days

                if idade > dias:
                    os.remove(filepath)
                    count += 1
                    logger.info(f"Arquivo removido (idade: {idade} dias): {filename}")

            if count > 0:
                logger.info(f"Limpeza concluída: {count} arquivo(s) removido(s)")
            else:
                logger.info("Nenhum arquivo antigo para remover")

        except Exception as e:
            logger.error(f"Erro ao limpar arquivos antigos: {e}")


def gerar_relatorio_completo(analyzer, output_dir: str = "output") -> dict:
    """
    Função auxiliar para gerar relatório completo

    Args:
        analyzer: Instância de RecargaAnalyzer
        output_dir: Diretório de saída

    Returns:
        Dict com caminhos dos arquivos gerados
    """
    try:
        generator = ReportGenerator(output_dir)

        # Gerar dados
        tabela_resumo = analyzer.gerar_tabela_resumo()
        tabela_codigos = analyzer.gerar_tabela_codigos()
        tabela_negadas = analyzer.gerar_tabela_negadas()
        ranking_negadas = analyzer.gerar_ranking_negadas()
        ranking_n2 = analyzer.gerar_ranking_n2()

        # Gerar gráfico 1: Ranking de todas as negadas
        grafico_negadas_path = generator.gerar_grafico_ranking(
            ranking_negadas,
            "Ranking de Origens - Todas as Recargas Negadas (Últimos 30min)",
            cor='#ffc107'  # Amarelo
        )

        # Gerar gráfico 2: Ranking específico de N2
        grafico_n2_path = generator.gerar_grafico_ranking(
            ranking_n2,
            "Ranking de Origens - Erros N2 (Últimos 30min)",
            cor='#dc3545'  # Vermelho
        )

        # Gerar Excel completo
        excel_path = generator.gerar_excel_completo(
            tabela_resumo,
            tabela_codigos,
            ranking_negadas,
            ranking_n2,
            tabela_negadas
        )

        return {
            'grafico_negadas': grafico_negadas_path,
            'grafico_n2': grafico_n2_path,
            'excel': excel_path,
            'ranking_negadas': ranking_negadas,
            'ranking_n2': ranking_n2
        }

    except Exception as e:
        logger.error(f"Erro ao gerar relatório completo: {e}")
        return {}
